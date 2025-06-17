import pandas as pd
import glob
import matplotlib.pyplot as plt
import openpyxl
import plotly.graph_objects as go
import plotly.express as px

class Tools:
    @staticmethod # static, because it doesn't need to access 
    #any class variables (no class instance necessary)
    
    def unmerge_excel(wb, filepath):
        sheet = wb['DTV']
        for merge_range in list(sheet.merged_cells):
            # Check if the merge range is in column A (min_col = 1)
            if merge_range.min_col == 1:
                sheet.unmerge_cells(str(merge_range))
        wb.save(filepath)
        
    def unmerge_excel_pre_2018(self, wb, filepath):
        sheet = wb['Daten - Données']
        for merge_range in list(sheet.merged_cells):
            # Check if the merge range is in column A (min_col = 1)
            if merge_range.min_col == 1:
                sheet.unmerge_cells(str(merge_range))
        wb.save(filepath)
        
    def find_row_with_value(self, sheet, value: str):
        value = int(value)
        for row in sheet.iter_rows(min_col=1, max_col=1):
            for cell in row:
                if cell.value is not None:
                    try:
                        cell_value = int(cell.value)  # Convert cell value to int
                        if cell_value == value:
                            return cell.row
                    except ValueError:
                        pass  # Ignore non-numeric cells
        return -1
    
    def __init__(self, folder_path):
        self.folder_path = folder_path
        
        
    def load_traffic_data(self, station_ID):
        print(f"Looking for files in: {self.folder_path}")  # Debugging line
        traffic_df = pd.DataFrame()
        
        # Find all Excel files in the folder
        excel_files = glob.glob(self.folder_path + "/*.xlsx")
        print(f"Excel files: {excel_files}")
        for file in excel_files:
            wb = openpyxl.load_workbook(file)
            Tools.unmerge_excel(wb, file)
            sheet = wb['DTV']
            row_index = self.find_row_with_value(sheet, station_ID)
            try:
                df = pd.read_excel(file, sheet_name='DTV', header=None, usecols='G:R', skiprows=row_index-1, nrows=1)
                print(f"Read Excel succesful: {file}") 
                year_df = pd.read_excel(file, sheet_name=1, header=None, usecols='C', skiprows=3, nrows=1)
                year = year_df.iloc[0, 0] if not year_df.empty else 'Unknown'
                # Create datetime index for each month
                months = pd.date_range(start=f'{year}-01', periods=len(df.columns), freq='ME')
                # Reshape data for time series
                df = pd.DataFrame({
                    'Date': months,
                    'Traffic': df.values.flatten()
                })
                df['Date'] = pd.to_datetime(df['Date'])
                # Append to the main dataframe
                traffic_df = pd.concat([df, traffic_df]) #, ignore_index=True)
            except Exception as e:
                print(f"Error reading file {file}: {e}")
                
        traffic_df = traffic_df.sort_values(by="Date") #.reset_index(drop=True)
        return traffic_df
    
    
    def load_traffic_data_pre_2018(self, station_ID):
        print(f"Looking for files in: {self.folder_path}")  # Debugging line
        traffic_df = pd.DataFrame()
        excel_files = glob.glob(self.folder_path + "/*.xlsx")
        print(f"Excel files found: {excel_files}")
        for file in excel_files:
            wb = openpyxl.load_workbook(file)
            sheet = wb['Daten - Données']
            self.unmerge_excel_pre_2018(wb, file)
                   
            row_index = self.find_row_with_value(sheet, station_ID)
            print(f"Row index: {row_index}")    
            try:
                year_df = pd.read_excel(file, sheet_name=1, header=None, usecols='C', skiprows=3, nrows=1)
                year = year_df.iloc[0, 0] if not year_df.empty else 'Unknown'
                print(f"Year: {year}, {type(year)}")
                # if year is 2008 until 2012, read different columns
                match year:
                    case (2008 | 2009 | 2010 | 2011 | 2012):
                        df = pd.read_excel(file, sheet_name='Daten - Données', header=None, skiprows=row_index-1, nrows=1)
                        df = df.iloc[:, 4:16]
                        print(f"Read Data of year {year} succesful: {df.head()}")
                    case (2013 | 2014 | 2015):
                        df = pd.read_excel(file, sheet_name='Daten - Données', header=None, usecols='G:R', skiprows=row_index-1, nrows=1)
                        print(f"Read Data of year {year} succesful: {df.head()}")
                    case (2016 | 2017):
                        df = pd.read_excel(file, sheet_name='Daten - Données', header=None, usecols='H:T', skiprows=row_index-1, nrows=1)
                        print(f"Read Data of year {year} succesful: {df.head()}")
                    case _:
                        print(f"Year {year} not in range")
                
                # Create datetime index for each month
                months = pd.date_range(start=f'{year}-01', periods=len(df.columns), freq='ME')
                # Reshape data for time series
                df = pd.DataFrame({
                    'Date': months,
                    'Traffic': df.values.flatten()
                })
                df['Date'] = pd.to_datetime(df['Date'])
                # Append to the main dataframe
                traffic_df = pd.concat([df, traffic_df]) #, ignore_index=True)
            except Exception as e:
                print(f"Error reading file {file}: {e}")
                
        traffic_df = traffic_df.sort_values(by="Date") #.reset_index(drop=True)
        return traffic_df
    
    @staticmethod
    def load_FR_air_data(folder_path):
        csv_files = glob.glob(folder_path + "/*.csv")

        air_df = pd.DataFrame()
        for file in csv_files:
            df = pd.read_csv(file, delimiter=';', skiprows=5, names=["Date", "Value"], parse_dates=["Date"])
            df['Date'] = pd.to_datetime(df['Date'])
            air_df = pd.concat([air_df, df], ignore_index=True)
        print(air_df.head())
        return air_df
    
    @staticmethod
    def load_AIGLE_air_data(folder_path):
        csv_files = glob.glob(folder_path + "/*.csv")

        air_df = pd.DataFrame()
        for file in csv_files:
            df = pd.read_csv(file, delimiter=';', skiprows=5, usecols=[0,1], names=["Date", "Value"], parse_dates=["Date"])
            df['Date'] = pd.to_datetime(df['Date'])
            air_df = pd.concat([air_df, df], ignore_index=True)
        print(air_df.head())
        return air_df
    @staticmethod
    def load_CAMORINO_air_data(folder_path):
        csv_file = f"{folder_path}/ID36_MAG.csv"
        air_df = pd.read_csv(csv_file, sep=";", skiprows=6, names=["Date", "O3", "NO2", "PM10", "PM2.5"], parse_dates=["Date"])
        air_df['Date'] = pd.to_datetime(air_df['Date'], format='%d.%m.%Y')
        return air_df
    
    
    @staticmethod
    def load_BASEL_air_data_NO2(file_path):
        air_df = pd.read_csv(file_path, delimiter=";", skiprows=5, usecols=[0,2], encoding="ISO-8859-1")
        air_df = air_df.rename(columns={"NO2 (Valeurs journalières moyennes  [µg/m³])": "Value (NO2 [µg/m³])"})
        air_df = air_df.dropna()
        air_df['Date'] = pd.to_datetime(air_df['Date'])
        return air_df
    def load_BASEL_air_data_PM10(file_path):
        air_df = pd.read_csv(file_path, delimiter=";", skiprows=5, usecols=[0,1], encoding="ISO-8859-1")
        air_df = air_df.rename(columns={"PM10 (Valeurs journalières moyennes  [µg/m³])": "Value (PM10 [µg/m³])"})
        air_df = air_df.dropna()
        air_df['Date'] = pd.to_datetime(air_df['Date'])
        return air_df
    def load_BASEL_air_data_PM25(file_path):
        air_df = pd.read_csv(file_path, delimiter=";", skiprows=5, usecols=[0,3], encoding="ISO-8859-1")
        air_df = air_df.rename(columns={"PM2.5 (Valeurs journalières moyennes  [µg/m³])": "Value (PM2.5 [µg/m³])"})
        air_df = air_df.dropna()
        air_df['Date'] = pd.to_datetime(air_df['Date'])
        return air_df

    @staticmethod
    def correlation2(csv1, csv2, method='pearson'):
        df1 = pd.read_csv(csv1, delimiter=',', parse_dates=['Date'])
        df2 = pd.read_csv(csv2, delimiter=',', parse_dates=['Date'])
        
        # Merge dataframes on 'Date'
        merged_df = pd.merge(df1, df2, on='Date', suffixes=('_df1', '_df2'))
        
        #remove column date
        merged_df = merged_df.drop(columns=['Date'])
        # Calculate correlation
        correlation = merged_df.corr(method=method)

        return correlation # return the correlation matrix in form of a dataframe
    
    @staticmethod
    def correlation3(csv1, csv2, csv3, method='pearson'):
        df1 = pd.read_csv(csv1, delimiter=',', parse_dates=['Date'])
        df2 = pd.read_csv(csv2, delimiter=',', parse_dates=['Date'])
        df3 = pd.read_csv(csv3, delimiter=',', parse_dates=['Date'])
        
        # Merge dataframes on 'Date'
        merged_df = pd.merge(df1, df2, on='Date', suffixes=('_df1', '_df2', '_df3'))
        merged_df = pd.merge(merged_df, df3, on='Date', suffixes=('_df1', '_df2', '_df3'))
        #remove column date
        merged_df = merged_df.drop(columns=['Date'])
        # Calculate correlation
        correlation = merged_df.corr(method=method)

        return correlation # return the correlation matrix in form of a dataframe
