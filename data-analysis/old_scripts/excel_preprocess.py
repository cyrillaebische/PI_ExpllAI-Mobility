import pandas as pd
import glob
import matplotlib.pyplot as plt
import openpyxl

def unmerge_excel(wb, filepath):
    sheet = wb['DTV']
    for merge_range in list(sheet.merged_cells):
        # Check if the merge range is in column A (min_col = 1)
        if merge_range.min_col == 1:
            sheet.unmerge_cells(str(merge_range))
    wb.save(filepath)
    

def find_row_with_value(sheet, value):
    for row in sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if cell.value == value:
                print(f'Row number for year: {cell.row}')
                return cell.row
            
def load_traffic_data(folder_path):
    traffic_df = pd.DataFrame()
    
    # Find all Excel files in the folder
    excel_files = glob.glob(folder_path + "/*.xlsx")
    for file in excel_files:
        wb = openpyxl.load_workbook(file)
        unmerge_excel(wb, file)
        sheet = wb['DTV']
        row_index = find_row_with_value(sheet, '175')
        try:
            df = pd.read_excel(file, sheet_name='DTV', header=None, usecols='G:R', skiprows=row_index-1, nrows=1)
            year_df = pd.read_excel(file, sheet_name=1, header=None, usecols='C', skiprows=3, nrows=1)
            year = year_df.iloc[0, 0] if not year_df.empty else 'Unknown'
            # Create datetime index for each month
            months = pd.date_range(start=f'{year}-01', periods=len(df.columns), freq='M')
            # Reshape data for time series
            df = pd.DataFrame({
                'Date': months,
                'Traffic Count': df.values.flatten()
            })
            
            # Append to the main dataframe
            traffic_df = pd.concat([df, traffic_df], ignore_index=True)
        except Exception as e:
            print(f"Error reading file {file}: {e}")
            
    return traffic_df

def plot_traffic_data(df):
    df.set_index('Date', inplace=True)
    df.plot(marker='o')
    plt.xlabel('Month')
    plt.ylabel('Traffic Count')
    plt.title('Monthly Traffic Count (24h-mean)')
    plt.xticks(rotation=45)
    plt.grid()
    plt.show()

traffic_data = load_traffic_data("data\ASTRA_Bulletins")
plot_traffic_data(traffic_data)

print(traffic_data)


        